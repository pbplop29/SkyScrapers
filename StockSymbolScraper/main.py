import openpyxl
import time
import requests
from bs4 import BeautifulSoup
wb = openpyxl.load_workbook("try.xlsx")
sheets = wb.sheetnames
#print(sheets)
sh = wb['Stock']
data = sh['F4'].value
#print(data)
s_row = 4
s_col = 6
c_list = []
while sh.cell(row=s_row, column= s_col).value != None:
    c_name = sh.cell(row=s_row, column=s_col).value
    c_list.append(c_name)
    s_row += 1
#print("Company name available in Database")
#[print('    ->', name) for name in c_list]
time.sleep(2)
for stock_symbol in c_list:
    url = 'https://www.tradingview.com/symbols/NSE-' + stock_symbol
    google = 'https://www.google.com/search?q=' + stock_symbol + '+share'
    response = requests.get(url)
    print(response)
    #print(google)
    print(url)
    soup = BeautifulSoup(page.read(), 'html.parser')
    #print(soup)
    #data = soup.find_all('div', attrs={'class': 'tv-symbol-price-quote__value js-symbol-last'})
    #data = soup.find_all('span', attrs={'class': 'tv-symbol-price-quote__value js-symbol-last'})
    div = soup.find("div", class_ = 'tv-symbol-price-quote__value js-symbol-last')
    price = div.span.text
    div.span.extract()
    title = div.get_text(strip=True)
    #print(title)
    print(price)
